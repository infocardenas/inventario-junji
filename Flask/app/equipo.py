from flask import request, flash, render_template, url_for, redirect, Blueprint, session, send_file
from db import mysql
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from cuentas import loguear_requerido, administrador_requerido
from werkzeug.utils import secure_filename
from cerberus import Validator
from MySQLdb import IntegrityError
from flask import jsonify
import os

equipo = Blueprint("equipo", __name__, template_folder="app/templates")


# envia datos al formulario y tabla de equipo CAMBIA FK_IDCODIGO_PROVEEDOR  
@equipo.route("/equipo")
@loguear_requerido
def Equipo():
    per_page = 7
    page = request.args.get('page', default=1, type=int)
    offset = (page - 1) * per_page

    cur = mysql.connection.cursor()

    # Total de registros
    cur.execute("SELECT COUNT(*) AS total FROM equipo")
    total = cur.fetchone()['total']
    total_pages = (total + per_page - 1) // per_page

    # Consulta paginada
    cur.execute("""
        SELECT *
        FROM super_equipo
        LIMIT %s OFFSET %s
    """, (per_page, offset))
    equipos = cur.fetchall()

    # Tipos de equipo
    cur.execute("SELECT * FROM tipo_equipo")
    tipo_equipo = cur.fetchall()

    # Unidades
    cur.execute("SELECT idUnidad, nombreUnidad FROM unidad")
    ubi_data = cur.fetchall()

    # Órdenes de compra
    cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
    ordenc_data = cur.fetchall()

    # Marcas
    cur.execute("SELECT * FROM marca_equipo")
    marcas = cur.fetchall()

    # Modelos por tipo
    modelos_por_tipo = {}
    for tipo in tipo_equipo:
        cur.execute("""
            SELECT me.*
            FROM modelo_equipo me
            INNER JOIN marca_tipo_equipo mte ON mte.idMarcaTipo = me.idMarca_Tipo_Equipo
            WHERE mte.idTipo_equipo = %s
        """, (tipo['idTipo_equipo'],))
        modelo_tipo = cur.fetchall()
        modelos_por_tipo[tipo['idTipo_equipo']] = modelo_tipo

    # Modelos con info adicional
    cur.execute("""
        SELECT me.*, te.nombreTipo_equipo, mae.nombreMarcaEquipo
        FROM modelo_equipo me
        INNER JOIN marca_tipo_equipo mte ON mte.idMarcaTipo = me.idMarca_Tipo_Equipo
        INNER JOIN tipo_equipo te ON te.idTipo_equipo = mte.idTipo_equipo
        INNER JOIN marca_equipo mae ON mae.idMarca_Equipo = mte.idMarca_Equipo
    """)
    modelo_equipo = cur.fetchall()

    # Estados
    cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo;")
    estados = cur.fetchall()

    # Provincias
    cur.execute("SELECT idProvincia, nombreProvincia FROM provincia")
    provincias = cur.fetchall()
    cur.close()

    marcas_llenadas = crear_lista_modelo_tipo_marca()

    return render_template(
        "Equipo/equipo.html",
        equipo=equipos,
        tipo_equipo=tipo_equipo,
        modelo_equipo_simple=modelo_equipo,
        marcas_equipo=marcas_llenadas,
        orden_compra=ordenc_data,
        Unidad=ubi_data,
        modelo_equipo=modelos_por_tipo,
        estado=estados,
        provincia=provincias,
        session=session,
        current_page=page,
        total_pages=total_pages
    )

def getPerPage():
    return 7 

def crear_lista_modelo_tipo_marca():
    cur = mysql.connection.cursor()

    # Obtener todas las marcas
    cur.execute("""
    SELECT *
    FROM marca_equipo
    """)
    marcas = cur.fetchall()
    marcas_llenadas = []

    for marca in marcas:
        tipos_llenados = []

        # Obtener los tipos de equipo asociados a esta marca a través de marca_tipo_equipo
        cur.execute("""
        SELECT te.*, mte.idMarcaTipo
        FROM tipo_equipo te
        INNER JOIN marca_tipo_equipo mte ON mte.idTipo_equipo = te.idTipo_equipo
        WHERE mte.idMarca_Equipo = %s
        """, (marca['idMarca_Equipo'],))
        tipos_equipo_asociados_marca = cur.fetchall()

        for tipo in tipos_equipo_asociados_marca:
            # Obtener los modelos asociados a este tipo y marca
            cur.execute("""
            SELECT me.*
            FROM modelo_equipo me
            INNER JOIN marca_tipo_equipo mte ON me.idMarca_Tipo_Equipo = mte.idMarcaTipo
            WHERE mte.idTipo_equipo = %s
            AND mte.idMarca_Equipo = %s
            """, (tipo['idTipo_equipo'], marca['idMarca_Equipo']))
            modelos_equipo_asociados_tipo = cur.fetchall()

            # Agregar los modelos al tipo
            tipo.update({'modelo_equipo': modelos_equipo_asociados_tipo})
            tipos_llenados.append(tipo)

        # Asociar los tipos llenados a la marca
        marca.update({'tipo_equipo': tipos_llenados})
        marcas_llenadas.append(marca)

    # Convertir la lista final en una tupla (opcional)
    return tuple(marcas_llenadas)

@equipo.route("/add_equipo", methods=["POST"])
@administrador_requerido
def add_equipo():
    if request.method == "POST":
        datos = {
            'codigo_inventario': request.form["codigo_inventario"].strip(),
            'numero_serie': request.form["numero_serie"].strip(),
            'observacion_equipo': request.form["observacion_equipo"].strip(),
            'codigoproveedor': request.form["codigoproveedor"].strip(),
            'mac': request.form["mac"].strip(),
            'imei': request.form["imei"].strip(),
            'numero': request.form["numero"].strip(),
            'codigo_Unidad': request.form["codigo_Unidad"].strip(),
            'nombre_orden_compra': request.form["nombre_orden_compra"].strip(),
            'idModelo_equipo': request.form["modelo_equipo"].strip(),
        }

        # Convertir cadenas vacías a None para los campos opcionales Error en los siguientes campos: - codigo_inventario: null value not allowed
        for key in ['mac', 'imei', 'numero', 'codigo_Unidad', 'nombre_orden_compra', 'idModelo_equipo', 'codigoproveedor']:
            if datos[key] == "":
                datos[key] = None

        # Convertir idModelo_equipo a entero si no es None
        if datos['idModelo_equipo']:
            datos['idModelo_equipo'] = int(datos['idModelo_equipo'])

        # Definir el esquema de validación
        schema = {
            'codigo_inventario': {'type': 'string','regex': '^[a-zA-Z0-9]+$'},
            'numero_serie': {'type': 'string', 'regex': '^[a-zA-Z0-9]+$'},
            'observacion_equipo': {'type': 'string', 'nullable': True},
            'codigoproveedor': {'type': 'string', 'regex': '^[a-zA-Z0-9]+$', 'nullable': True},
            'mac': {'type': 'string', 'regex': r'^([0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}$', 'nullable': True},
            'imei': {'type': 'string', 'regex': '^[0-9]+$', 'nullable': True},
            'numero': {'type': 'string', 'regex': '^[0-9]+$', 'nullable': True},
            'codigo_Unidad': {'type': 'string', 'nullable': True},
            'nombre_orden_compra': {'type': 'string', 'nullable': True},
            'idModelo_equipo': {'type': 'integer', 'nullable': True},
        }

        # Validar los datos usando Cerberus
        v = Validator(schema)
        if not v.validate(datos):
            errores = v.errors
            mensaje_error = "Error en los siguientes campos:\n"
            for campo, error in errores.items():
                mensaje_error += f"- {campo}: {', '.join(error)}\n"
            flash(mensaje_error, 'warning')
            return redirect(url_for("equipo.Equipo"))

        # 🔍 **Conexión con la base de datos**
        try:
            cur = mysql.connection.cursor()

            # 🛠 **Verificar duplicados antes de la inserción**

            # ✅ Verificar si el código de inventario ya existe
            if datos['codigo_inventario']:
                cur.execute("SELECT idEquipo FROM equipo WHERE Cod_inventarioEquipo = %s", (datos['codigo_inventario'],))
                equipo_existente = cur.fetchone()
                if equipo_existente:
                    # Excepción: permitir si el equipo existente tiene incidencia con estado 'equipo cambiado'
                    cur.execute("""
                        SELECT 1 FROM incidencia
                        WHERE idEquipo = %s AND estadoIncidencia = 'equipo cambiado'
                        """, (equipo_existente['idEquipo'],))
                    incidencia_cambiado = cur.fetchone()
                    if not incidencia_cambiado:
                        flash("El código de inventario ya está en uso", 'warning')
                        return redirect(url_for("equipo.Equipo"))

            # ✅ Verificar si el número de serie ya existe
            if datos['numero_serie']:
                cur.execute("SELECT idEquipo FROM equipo WHERE Num_serieEquipo = %s", (datos['numero_serie'],))
                if cur.fetchone():
                    flash("El número de serie ya existe", 'warning')
                    return redirect(url_for("equipo.Equipo"))

            # ✅ Verificar si el código de proveedor ya existe
            if datos['codigoproveedor']:
                cur.execute("SELECT idEquipo FROM equipo WHERE codigoproveedor_equipo = %s", (datos['codigoproveedor'],))
                if cur.fetchone():
                    flash("El código de proveedor ya está en uso", 'warning')
                    return redirect(url_for("equipo.Equipo"))

            # ✅ Verificar si el número telefónico ya existe
            if datos['numero']:
                cur.execute("SELECT idEquipo FROM equipo WHERE numerotelefonicoEquipo = %s", (datos['numero'],))
                if cur.fetchone():
                    flash("El número telefónico ya está en uso", 'warning')
                    return redirect(url_for("equipo.Equipo"))

            # ✅ Verificar existencia del modelo
            if datos['idModelo_equipo']:
                cur.execute("SELECT COUNT(*) AS count FROM modelo_equipo WHERE idModelo_Equipo = %s", 
                            (datos['idModelo_equipo'],))
                modelo_existe = cur.fetchone()['count']
                if modelo_existe == 0:
                    flash("El modelo seleccionado no existe", "warning")
                    return redirect(url_for("equipo.Equipo"))

            # 🛠 **Insertar equipo en la base de datos**
            cur.execute(
                """ INSERT INTO equipo (
                    Cod_inventarioEquipo, 
                    Num_serieEquipo, 
                    ObservacionEquipo,
                    codigoproveedor_equipo,
                    macEquipo,
                    imeiEquipo, 
                    numerotelefonicoEquipo, 
                    idEstado_Equipo, 
                    idUnidad, 
                    idOrden_compra, 
                    idModelo_equipo) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (
                    datos['codigo_inventario'] or None,
                    datos['numero_serie'] or None,
                    datos['observacion_equipo'] or None,
                    datos['codigoproveedor'] or None,
                    datos['mac'] or None,
                    datos['imei'] or None,
                    datos['numero'] or None,
                    1,
                    datos['codigo_Unidad'] or None,
                    datos['nombre_orden_compra'] or None,
                    datos['idModelo_equipo'] or None,
                ),
            )
            mysql.connection.commit()
            flash("Equipo agregado correctamente", 'success')
            return redirect(url_for("equipo.Equipo"))

        except Exception as e:
            mysql.connection.rollback()
            flash(f"Error al crear el equipo: {str(e)}", "warning")
            return redirect(url_for('equipo.Equipo'))


@equipo.route("/update_equipo/<id>", methods=["POST"])
@administrador_requerido
def update_equipo(id):
    if "user" not in session:
        flash("No estás autorizado", "danger")
        return redirect("/ingresar")
    try:
        cur = mysql.connection.cursor()

        # Obtener los datos enviados por el formulario (envío tradicional)
        datos = {
            'codigo_inventario': request.form.get("codigo_inventario", "").strip(),
            'numero_serie': request.form.get("numero_serie", "").strip(),
            'observacion_equipo': request.form.get("observacion_equipo", "").strip(),
            'codigoproveedor': request.form.get("codigoproveedor", "").strip(),
            'mac': request.form.get("mac", "").strip(),
            'imei': request.form.get("imei", "").strip(),
            'numero': request.form.get("numero", "").strip(),
            'codigo_Unidad': request.form.get("codigo_Unidad", "").strip(),
            'nombre_orden_compra': request.form.get("nombre_orden_compra", "").strip(),
            # Recibimos los id en lugar de nombres para estos campos:
            'tipo': request.form.get("tipo", "").strip(),
            'marca': request.form.get("marca", "").strip(),
            'modelo': request.form.get("modelo", "").strip(),
            # Agregamos la captura del estado (idEstado_equipo)
            'estado_equipo': request.form.get("estado_equipo", "").strip()
        }

        # Convertir cadenas vacías a None para campos opcionales
        for key in ['mac', 'imei', 'numero', 'codigo_Unidad', 'nombre_orden_compra', 'codigoproveedor']:
            if datos[key] == "":
                datos[key] = None

        # Definir el esquema de validación con Cerberus
        schema = {
            'codigo_inventario': {'type': 'string', 'regex': '^[a-zA-Z0-9]+$'},
            'numero_serie': {'type': 'string', 'regex': '^[a-zA-Z0-9]+$'},
            'observacion_equipo': {'type': 'string', 'nullable': True},
            'codigoproveedor': {'type': 'string', 'regex': '^[a-zA-Z0-9]+$', 'nullable': True},
            'mac': {'type': 'string', 'regex': '^([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})$', 'nullable': True},
            'imei': {'type': 'string', 'regex': '^[0-9]+$', 'nullable': True},
            'numero': {'type': 'string', 'regex': '^[0-9]+$', 'nullable': True},
            'codigo_Unidad': {'type': 'string', 'nullable': True},
            'nombre_orden_compra': {'type': 'string', 'nullable': True},
            # Campos de tipo, marca y modelo se esperan como id
            'tipo': {'type': 'string', 'nullable': False},
            'marca': {'type': 'string', 'nullable': False},
            'modelo': {'type': 'string', 'nullable': False},
            # Validamos también estado_equipo (id)
            'estado_equipo': {'type': 'string', 'regex': '^[0-9]+$', 'nullable': False},
        }

        v = Validator(schema)
        if not v.validate(datos):
            errores = v.errors
            mensaje_error = "Error en los siguientes campos:\n"
            for campo, error in errores.items():
                mensaje_error += f"- {campo}: {', '.join(error)}\n"
            flash(mensaje_error, 'warning')
            return redirect(url_for("equipo.Equipo"))

        # Verificar si el código de inventario ya existe (excepto para el mismo equipo)

        # Verificar si el número de serie ya existe (excepto para el mismo equipo)
        cur.execute("""
            SELECT idEquipo FROM equipo 
            WHERE Num_serieEquipo = %s AND idEquipo != %s
        """, (datos['numero_serie'], id))
        if cur.fetchone():
            flash("El número de serie ya está en uso", 'warning')
            return redirect(url_for("equipo.Equipo"))

        # Obtener ID del modelo utilizando el id recibido (no el nombre)
        try:
            modelo_id = int(datos['modelo'])
        except ValueError:
            flash("El modelo seleccionado es inválido", 'warning')
            return redirect(url_for("equipo.Equipo"))

        cur.execute("SELECT idModelo_Equipo FROM modelo_equipo WHERE idModelo_Equipo = %s", (modelo_id,))
        modelo = cur.fetchone()
        if not modelo:
            flash("El modelo seleccionado no existe", 'warning')
            return redirect(url_for("equipo.Equipo"))

        # Obtener ID de estado_equipo
        try:
            estado_equipo_id = int(datos['estado_equipo'])
        except ValueError:
            flash("El estado seleccionado es inválido", 'warning')
            return redirect(url_for("equipo.Equipo"))

        # Verificar que el estado exista en la tabla estado_equipo (opcional pero recomendable)
        cur.execute("SELECT idEstado_equipo FROM estado_equipo WHERE idEstado_equipo = %s", (estado_equipo_id,))
        estado_valido = cur.fetchone()
        if not estado_valido:
            flash("El estado seleccionado no existe", 'warning')
            return redirect(url_for("equipo.Equipo"))

        # Actualizar equipo (incluyendo el estado)
        cur.execute("""
            UPDATE equipo
            SET Cod_inventarioEquipo = %s, 
                Num_serieEquipo = %s, 
                ObservacionEquipo = %s, 
                codigoproveedor_equipo = %s, 
                macEquipo = %s, 
                imeiEquipo = %s, 
                numerotelefonicoEquipo = %s, 
                idUnidad = %s, 
                idOrden_compra = %s, 
                idModelo_equipo = %s,
                idEstado_equipo = %s
            WHERE idEquipo = %s
        """, (
            datos['codigo_inventario'], 
            datos['numero_serie'], 
            datos['observacion_equipo'], 
            datos['codigoproveedor'], 
            datos['mac'], 
            datos['imei'], 
            datos['numero'], 
            datos['codigo_Unidad'], 
            datos['nombre_orden_compra'], 
            modelo['idModelo_Equipo'],
            estado_equipo_id,  # <--- Se actualiza aquí
            id
        ))

        mysql.connection.commit()
        flash("Equipo editado correctamente", 'success')
        return redirect(url_for("equipo.Equipo"))

    except IntegrityError as e:
        mysql.connection.rollback()
        mensaje_error = str(e)
        if "Duplicate entry" in mensaje_error:
            if "Cod_inventarioEquipo" in mensaje_error:
                flash("El código de inventario ya existe", 'warning')
            elif "Num_serieEquipo" in mensaje_error:
                flash("El número de serie ya existe", 'warning')
            else:
                flash("Error de duplicación en la base de datos", 'warning')
        else:
            flash("Error de integridad en la base de datos", 'danger')
        return redirect(url_for("equipo.Equipo"))

    except Exception as error:
        mysql.connection.rollback()
        flash(f"Error al actualizar el equipo: {str(error)}", 'danger')
        return redirect(url_for("equipo.Equipo"))





@equipo.route("/delete_equipo/<id>", methods=["POST", "GET"])
@administrador_requerido
def delete_equipo(id):
    if "user" not in session:
        flash("No estás autorizado", "warning")
        return redirect("/ingresar")

    try:
        cur = mysql.connection.cursor()

        # **Validar si el equipo tiene dependencias en otras tablas**
        dependencias_queries = {
            "asignaciones": "SELECT COUNT(*) AS count FROM asignacion WHERE idAsignacion IN (SELECT idAsignacion FROM equipo_asignacion WHERE idEquipo = %s)",
            "equipo_asignacion": "SELECT COUNT(*) AS count FROM equipo_asignacion WHERE idEquipo = %s",
            "traslaciones": "SELECT COUNT(*) AS count FROM traslacion WHERE idEquipo = %s",
            "incidencias": "SELECT COUNT(*) AS count FROM incidencia WHERE idEquipo = %s"
        }

        dependencias = {}
        for key, query in dependencias_queries.items():
            cur.execute(query, (id,))
            dependencias[key] = cur.fetchone()["count"]

        # **Si hay dependencias, NO eliminar y mostrar un mensaje claro**
        if any(count > 0 for count in dependencias.values()):
            mensaje = "No se puede eliminar el equipo porque tiene registros relacionados en:\n"
            for key, count in dependencias.items():
                if count > 0:
                    mensaje += f"- {key.replace('_', ' ').capitalize()}: {count} registros.\n"
            flash(mensaje, "danger")
            return redirect(url_for("equipo.Equipo"))

        # **Si no tiene dependencias, proceder con la eliminación**
        cur.execute("DELETE FROM equipo_asignacion WHERE idEquipo = %s", (id,))
        cur.execute("DELETE FROM traslacion WHERE idEquipo = %s", (id,))
        cur.execute("DELETE FROM incidencia WHERE idEquipo = %s", (id,))
        cur.execute("DELETE FROM equipo WHERE idEquipo = %s", (id,))
        mysql.connection.commit()

        flash("Equipo eliminado correctamente", "success")

    except Exception as e:
        mysql.connection.rollback()
        error_message = f"Error al eliminar el equipo: {str(e)}"
        print(error_message)  # Registrar el error en la terminal
        flash(error_message, "danger")

    finally:
        try:
            cur.close()
        except Exception as close_error:
            print(f"Error al cerrar el cursor: {str(close_error)}")

        return redirect(url_for("equipo.Equipo"))

@equipo.route("/mostrar_asociados_traslado/<idTraslado>")
@loguear_requerido
def mostrar_asociados_traslado(idTraslado):
    page = 1
    perpage = 200
    offset = (page - 1) * perpage
    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) FROM equipo")
    total = cur.fetchone()
    total = int(str(total).split(":")[1].split("}")[0])
    cur.execute(
        """ 
    SELECT e.idEquipo, e.Cod_inventarioEquipo, e.Num_serieEquipo, e.ObservacionEquipo, 
    e.codigoproveedor_equipo, e.macEquipo, e.imeiEquipo, e.numerotelefonicoEquipo,
    e.idEstado_Equipo, e.idUnidad, e.idOrden_compra, e.idModelo_equipo,te.idTipo_equipo, 
    te.nombreTipo_Equipo, ee.idEstado_equipo, ee.nombreEstado_equipo, u.idUnidad, u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
    moe.idModelo_equipo, moe.nombreModeloequipo
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idMarca_Tipo_Equipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    INNER JOIN traslacion tras on tras.idEquipo = e.idEquipo
    WHERE tras.idTraslado = %s
    LIMIT %s OFFSET %s
    """,
        (idTraslado, perpage, offset),
    )
    data = cur.fetchall()
    cur.execute("SELECT * FROM tipo_equipo")
    tipoe_data = cur.fetchall()
    cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo")
    estadoe_data = cur.fetchall()
    cur.execute("SELECT idUnidad, nombreUnidad FROM unidad")
    ubi_data = cur.fetchall()
    cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
    ordenc_data = cur.fetchall()
    cur.execute("SELECT idModelo_Equipo, nombreModeloequipo FROM modelo_equipo")
    modeloe_data = cur.fetchall()

    return render_template(
        "Equipo/equipo.html",
        equipo=data,
        tipo_equipo=tipoe_data,
        estado_equipo=estadoe_data,
        orden_compra=ordenc_data,
        Unidad=ubi_data,
        modelo_equipo=modeloe_data,
        page=page,
        lastpage=False,
    )


@equipo.route("/mostrar_asociados_unidad/<idUnidad>")
@equipo.route("/mostrar_asociados_unidad/<idUnidad>/<page>")
@loguear_requerido
def mostrar_asociados_unidad(idUnidad, page=1):
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    page = int(page)
    page = 1
    perpage = 200  
    offset = (page - 1) * perpage
    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) FROM equipo")
    total = cur.fetchone()
    total = int(str(total).split(":")[1].split("}")[0])
    cur.execute(
        """ 
    SELECT e.idEquipo, e.Cod_inventarioEquipo, e.Num_serieEquipo, e.ObservacionEquipo, 
    e.codigoproveedor_equipo, e.macEquipo, e.imeiEquipo, e.numerotelefonicoEquipo,
    e.idEstado_Equipo, e.idUnidad, e.idOrden_compra, e.idModelo_equipo,te.idTipo_equipo, 
    te.nombreTipo_Equipo, ee.idEstado_equipo, ee.nombreEstado_equipo, 
    u.idUnidad, u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
    moe.idModelo_equipo, moe.nombreModeloequipo
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    WHERE e.idUnidad = %s
    LIMIT %s OFFSET %s
    """,
        (
            idUnidad,
            perpage,
            offset,
        ),
    )
    data = cur.fetchall()
    cur.execute("SELECT * FROM tipo_equipo")
    tipoe_data = cur.fetchall()
    cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo")
    estadoe_data = cur.fetchall()
    cur.execute("SELECT idUnidad, nombreUnidad FROM unidad")
    ubi_data = cur.fetchall()
    cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
    ordenc_data = cur.fetchall()
    cur.execute("SELECT idModelo_Equipo, nombreModeloequipo FROM modelo_equipo")
    modeloe_data = cur.fetchall()

    return render_template(
        "Equipo/equipo.html",
        equipo=data,
        tipo_equipo=tipoe_data,
        estado_equipo=estadoe_data,
        orden_compra=ordenc_data,
        Unidad=ubi_data,
        modelo_equipo=modeloe_data,
        page=page,
        lastpage=False,
    )

@equipo.route("/mostrar_asociados_funcionario/<rutFuncionario>")
@equipo.route("/mostrar_asociados_funcionario/<rutFuncionario>/<page>")
@loguear_requerido
def mostrar_asociados_funcionario(rutFuncionario, page=1):
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM tipo_equipo")
    tipoe_data = cur.fetchall()
    cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo")
    estadoe_data = cur.fetchall()
    cur.execute("SELECT idUnidad, nombreUnidad FROM unidad")
    ubi_data = cur.fetchall()
    cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
    ordenc_data = cur.fetchall()
    cur.execute("SELECT idModelo_Equipo, nombreModeloequipo FROM modelo_equipo")
    modeloe_data = cur.fetchall()

    page = int(page)
    page = 1
    perpage = 200 #porque ningun funcionario tendra tantos 
                    #equipos que la paginacion sea nesesaria
    offset = (page - 1) * perpage
    cur.execute("SELECT COUNT(*) FROM equipo")
    total = cur.fetchone()
    total = int(str(total).split(":")[1].split("}")[0])
    #encontar la fecha de la ultima asignacion
    cur.execute("""
                SELECT *
                FROM asignacion a
                WHERE a.rutFuncionario = %s
                AND a.ActivoAsignacion = 1
                ORDER BY a.fecha_inicioAsignacion DESC
                """, (rutFuncionario,))
    asignaciones = cur.fetchall()

    if(len(asignaciones) == 0):
        
        return render_template(
            "Equipo/equipo.html",
            equipo=(),
            tipo_equipo=tipoe_data,
            estado_equipo=estadoe_data,
            orden_compra=ordenc_data,
            Unidad=ubi_data,
            modelo_equipo=modeloe_data,
            page=page,
            lastpage=False,
        )
    ultimaAsignacion = asignaciones[0]

    cur.execute(""" 
    SELECT *
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    INNER JOIN equipo_asignacion ea on ea.idEquipo = e.idEquipo
    INNER JOIN asignacion a on a.idAsignacion = ea.idAsignacion
    INNER JOIN funcionario f on f.rutFuncionario = a.rutFuncionario
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    WHERE a.idAsignacion = %s AND f.rutFuncionario = %s
    LIMIT %s OFFSET %s
    """,
        (
            ultimaAsignacion['idAsignacion'],
            rutFuncionario,
            perpage,
            offset,
        ),
    )
    data = cur.fetchall()
    return render_template(
        "Equipo/equipo.html",
        equipo=data,
        tipo_equipo=tipoe_data,
        estado_equipo=estadoe_data,
        orden_compra=ordenc_data,
        Unidad=ubi_data,
        modelo_equipo=modeloe_data,
        page=page,
        lastpage=False,
    )


@equipo.route("/equipo_detalles/<idEquipo>")
@loguear_requerido
def equipo_detalles(idEquipo):
    cur = mysql.connection.cursor()
    cur.execute("""
            SELECT i.fechaIncidencia as fecha, i.idIncidencia as id,
                "Incidencia" as evento, i.observacionIncidencia as observacion,
                i.nombreIncidencia as nombre
            FROM incidencia i
            WHERE i.idEquipo = %s
            UNION ALL
            SELECT traslado.fechaTraslado, traslado.idTraslado, "Traslado",
                "Nombre", "observacion"
            FROM traslado, traslacion
            WHERE traslacion.idTraslado = traslado.idTraslado AND traslacion.idEquipo = %s
            UNION ALL
            SELECT a.fecha_inicioAsignacion, a.idAsignacion, "Asignacion",
                a.ObservacionAsignacion, f.nombreFuncionario 
            FROM asignacion a
            INNER JOIN funcionario f on f.rutFuncionario = a.rutFuncionario
            INNER JOIN equipo_asignacion ea on a.idAsignacion = ea.idAsignacion
            WHERE ea.idEquipo = %s
            UNION ALL
            SELECT d.fechaDevolucion, ea.idAsignacion, "Devolucion",
                d.observacionDevolucion, f.nombreFuncionario
            FROM devolucion d
            INNER JOIN equipo_asignacion ea on d.idEquipoAsignacion = ea.idEquipoAsignacion
            INNER JOIN asignacion a on ea.idAsignacion = a.idAsignacion
            INNER JOIN funcionario f on f.rutFuncionario = a.rutFuncionario
            WHERE ea.idEquipo = %s
            ORDER BY fecha DESC
            """, (idEquipo, idEquipo, idEquipo, idEquipo))

    data_eventos = cur.fetchall()
    cur.execute("""
    SELECT *
    FROM super_equipo se
    WHERE se.idEquipo = %s
    """, (idEquipo,)
    )
    data_equipo = cur.fetchone()
    #revisar si este equipo esta asignado a un funcionario
    if data_equipo['nombreFuncionario'] != '':
        cur.execute("""
        SELECT *
        FROM funcionario f
        INNER JOIN unidad u ON u.idUnidad = f.idUnidad
        WHERE f.rutFuncionario = %s
        """, (data_equipo['rutFuncionario'],))
        funcionario = cur.fetchone()
    else:
        funcionario = None

    return render_template(
        "Equipo/equipo_detalles.html", 
        equipo=data_equipo, 
        eventos=data_eventos, 
        funcionario=funcionario
        )

@equipo.route("/equipo/importar_excel", methods=["POST"])
@administrador_requerido
def importar_excel():
    #Nesesito el excel para ver el formato
    file = request.files["file"]
    path = ""
    safename = secure_filename(file.filename)
    file.save(os.path.join(path, safename))
    wb = load_workbook(os.path.join(path, safename))
    ws = wb['Equipo']
    importar_equipo(col_codigo_inventario='A', 
                    col_n_serie='B', 
                    col_codigo_proveedor='C', 
                    col_mac='D', 
                    col_imei='E', 
                    col_telefono='F', 
                    col_idUnidad='G', 
                    col_modelo='H', 
                    col_tipo_equipo='I',
                    col_marca='J', 
                    col_id_orden_compra='P',
                    col_nombre_orden_compra='K', 
                    col_fecha_inicio_compra='O',
                    col_fecha_fin_compra='L',
                    col_tipo_adquisicion='M',
                    col_proveedor='N',
                    
                    worksheet=ws)
    return redirect("/equipo")

@equipo.route("/equipo/importar_excel/unidad", methods=["POST"])
@administrador_requerido
def importar_excel_unidad():
    #Nesesito el excel para ver el formato
    file = request.files["file"]
    path = ""
    safename = secure_filename(file.filename)
    file.save(os.path.join(path, safename))
    wb = load_workbook(os.path.join(path, safename))
    ws = wb['Unidad']
    importar_unidad(col_comuna='E',
                    col_modalidad='F', 
                    col_codigo='A', 
                    col_contacto='C', 
                    col_nombre='B', 
                    col_direccion='D', 
                    worksheet=ws)
    return redirect("/equipo")
    


def importar_equipo(col_codigo_inventario, col_n_serie, col_codigo_proveedor, 
                    col_mac, col_imei, col_telefono, col_idUnidad, col_modelo,
                    col_tipo_equipo, col_marca, col_id_orden_compra,
                    col_nombre_orden_compra, col_fecha_inicio_compra,
                    col_fecha_fin_compra, col_tipo_adquisicion,
                    col_proveedor, worksheet):
    print("importar equipos")
    

    for row in range(2, worksheet.max_row+1):
        codigo_inventario = worksheet[col_codigo_inventario + str(row)].value
        n_serie = worksheet[col_n_serie + str(row)].value
        codigo_proveedor = worksheet[col_codigo_proveedor + str(row)].value
        mac = worksheet[col_mac + str(row)].value
        imei = worksheet[col_imei + str(row)].value
        telefono = worksheet[col_telefono + str(row)].value
        idUnidad = worksheet[col_idUnidad + str(row)].value
        modelo = worksheet[col_modelo + str(row)].value
        nombre_tipo = worksheet[col_tipo_equipo + str(row)].value
        nombre_marca = worksheet[col_marca + str(row)].value
        id_orden_compra = worksheet[col_id_orden_compra + str(row)].value
        nombre_orden_compra = worksheet[col_nombre_orden_compra + str(row)].value
        fecha_inicio_compra = worksheet[col_fecha_inicio_compra + str(row)].value
        fecha_fin_compra = worksheet[col_fecha_fin_compra + str(row)].value
        nombre_tipo_adquisicion = worksheet[col_tipo_adquisicion + str(row)].value
        nombre_proveedor = worksheet[col_proveedor + str(row)].value

        cur = mysql.connection.cursor()

        #nombre orden de compra
        if(nombre_orden_compra != None):
            cur.execute("""
            SELECT *
            FROM orden_compra oo
            WHERE LOWER(oo.idOrden_compra) LIKE LOWER(%s)
            """, (id_orden_compra,))
            tupla_orden_compra = cur.fetchall()
            print('tupla_orden_compra')
            print(tupla_orden_compra)
            print(len(tupla_orden_compra))
            if(len(tupla_orden_compra) == 0):
                #revisar proveedor
                cur.execute("""
                SELECT *
                FROM tipo_adquisicion ta
                WHERE LOWER(ta.nombre_tipo_adquisicion) LIKE LOWER(%s)
                """, (nombre_tipo_adquisicion,))
                tupla_adquisicion = cur.fetchall()
                idTipo_adquisicion = ""
                if(len(tupla_adquisicion) == 0):
                    cur.execute("""
                    INSERT INTO tipo_adquisicion(nombre_tipo_adquisicion)
                    VALUES (%s)
                    """, (nombre_tipo_adquisicion,))
                    mysql.connection.commit()
                    idTipo_adquisicion = cur.lastrowid
                else:
                    idTipo_adquisicion = tupla_adquisicion[0]['idTipo_adquisicion']

                #revisar tipo adquisicion

                cur.execute("""
                SELECT *
                FROM proveedor p
                WHERE LOWER(p.nombreProveedor) = LOWER(%s)
                """, (nombre_proveedor,))
                tupla_proveedor = cur.fetchall()
                idProveedor = ""
                if(len(tupla_proveedor) == 0):
                    cur.execute("""
                    INSERT INTO proveedor
                    (nombreProveedor)
                    VALUES (%s)
                    """, (nombre_proveedor,))
                    mysql.connection.commit()
                    idProveedor = cur.lastrowid
                else:
                    idProveedor = tupla_proveedor[0]['idProveedor']


                print("id_orden_compra")
                print(id_orden_compra)
                #crear orden de compra
                cur.execute("""
                INSERT INTO orden_compra
                (idOrden_compra ,nombreOrden_compra, fechacompraOrden_compra,
                fechafin_ORDEN_COMPRA, idTipo_adquisicion,
                            idProveedor)
                VALUES(%s, %s, %s,
                        %s, %s,
                            %s)
                            """,(id_orden_compra, nombre_orden_compra, fecha_inicio_compra,
                                fecha_fin_compra, idTipo_adquisicion, idProveedor))
                mysql.connection.commit()
            else:
                id_orden_compra = tupla_orden_compra[0]['idOrden_compra']
        else:
            id_orden_compra = None
        #buscar modelo con un nombre igual

        print("id_orden_equipo")
        print(id_orden_compra)

        cur.execute("""
        SELECT me.idModelo_Equipo
        FROM modelo_equipo me
        WHERE LOWER(me.nombreModeloequipo) = LOWER(%s)
                    """, (modelo,))
        
        id_modelo_equipo = ""
        tupla_modelo = cur.fetchall()
        if(len(tupla_modelo) == 0):
            #insertar el modelo

            Ids = encontrar_o_crear_tipo_equipo(nombre_tipo, cur, nombre_marca)
            id_tipo_equipo = Ids[0]
            id_marca_equipo = Ids[1]

            print("insert modelo")
            cur.execute("""
            INSERT INTO modelo_equipo
                (nombreModeloequipo,
                idTipo_Equipo,
                idMarca_Equipo)
            VALUES (%s, %s, %s)
                """, (modelo, id_tipo_equipo, id_marca_equipo))
            mysql.connection.commit()
            id_modelo_equipo = cur.lastrowid
        else:
            id_modelo_equipo = tupla_modelo[0]['idModelo_Equipo']

        #buscar id de SIN ASIGNAR
        cur.execute("""
        SELECT idEstado_equipo
        FROM estado_equipo
        WHERE LOWER(nombreEstado_equipo) LIKE LOWER(%s)
        """, ("SIN ASIGNAR",))
        idEstado_equipo = cur.fetchone()['idEstado_equipo']

        print("insertar equipo")
        cur.execute("""
        INSERT INTO equipo 
                (Cod_inventarioEquipo,
                Num_serieEquipo,
                ObservacionEquipo,
                codigoproveedor_equipo,
                macEquipo,
                imeiEquipo,
                numerotelefonicoEquipo,
                idEstado_equipo,
                idUnidad,
                idOrden_compra,
                idModelo_Equipo)

            VALUES (%s, 
                    %s, 
                    %s, 
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s)
                """, (codigo_inventario,
                      n_serie,
                      None,
                    codigo_proveedor,
                    mac,
                    imei,
                    telefono,
                    idEstado_equipo,
                    idUnidad,
                    id_orden_compra,
                    id_modelo_equipo
                      ))
        mysql.connection.commit()
        flash("equipos importados")
        pass
    pass 

def encontrar_o_crear_tipo_equipo(nombreTipoEquipo, cur, nombre_marca):
    cur.execute("""
    SELECT idTipo_equipo
    FROM tipo_equipo te
    WHERE te.nombreTipo_equipo = %s
                """, (nombreTipoEquipo,))
    tupla_tipo_equipos = cur.fetchall()
    print("tupla_tipo_equipo")
    print(tupla_tipo_equipos)
    id_tipo_equipo = ""
    if(len(tupla_tipo_equipos) == 0):
        cur.execute("""
        INSERT INTO tipo_equipo
            (nombreTipo_Equipo)
        VALUES (%s)
        """, (nombreTipoEquipo,))
        mysql.connection.commit()
        id_tipo_equipo = cur.lastrowid
    else:
        id_tipo_equipo = tupla_tipo_equipos['idTipo_equipo']
    
    #revisar si existe una marca con el nombre enviado como argumento
    cur.execute("""
    SELECT idMarca_Equipo
    FROM marca_equipo me
    WHERE LOWER(me.nombreMarcaEquipo) LIKE LOWER(%s)
    """, (nombre_marca,))
    Marca = cur.fetchall()
    id_marca_equipo = ""
    if(len(Marca) == 0):
        #crear marca con el nombre
        cur.execute("""
        INSERT INTO marca_equipo
        (nombreMarcaEquipo)
        VALUES (%s)
                """, (nombre_marca,))
        mysql.connection.commit()
        id_marca_equipo = cur.lastrowid
    else:
        id_marca_equipo = Marca[0]['idMarca_Equipo']
    #verificar que el tipo y la marca esten asociados de lo contrario crear asociacion
    cur.execute("""
    SELECT *
    FROM marca_tipo_equipo mte
    WHERE mte.idMarca_Equipo = %s AND mte.idTipo_equipo = %s
                """, (id_marca_equipo, id_tipo_equipo))
    Marca_tipo_equipo = cur.fetchall()

    if(len(Marca_tipo_equipo) == 0):
        #agregar la relacion
        cur.execute("""
        INSERT INTO marca_tipo_equipo 
        (idMarca_Equipo, idTipo_equipo)
        VALUES (%s, %s)
        """, (id_marca_equipo, id_tipo_equipo,))
        mysql.connection.commit()
    
    return (id_tipo_equipo, id_marca_equipo)


def importar_unidad(col_comuna, col_modalidad, col_codigo, 
                    col_contacto, col_nombre, col_direccion, worksheet):
    print("importar unidad" + str(worksheet.max_row))
    for row in range(2, worksheet.max_row+1):
        nombreComuna = worksheet[col_comuna + str(row)].value
        nombreModalidad = worksheet[col_modalidad + str(row)].value
        codigoUnidad = worksheet[col_codigo + str(row)].value
        contactoUnidad = worksheet[col_contacto + str(row)].value
        nombreUnidad = worksheet[col_nombre + str(row)].value
        direccionUnidad = worksheet[col_direccion + str(row)].value

        cur = mysql.connection.cursor()
        print("importar_unidad")

        cur.execute("""
            SELECT idComuna
            FROM comuna c
            WHERE LOWER(c.nombreComuna) LIKE LOWER(%s)
                    """, (nombreComuna,))
        Comuna = cur.fetchone()
        print(Comuna)
        cur.execute("""
            SELECT idModalidad
            FROM modalidad m
            WHERE LOWER(m.nombreModalidad) LIKE LOWER(%s)
                    """, (nombreModalidad,))
        Modalidad = cur.fetchone()

        try:
            cur.execute("""
            INSERT INTO unidad
                        (idUnidad, nombreUnidad, contactoUnidad, direccionUnidad, idComuna, idModalidad)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """, (codigoUnidad, nombreUnidad, contactoUnidad, direccionUnidad, 
                        Comuna['idComuna'], Modalidad['idModalidad']))
            mysql.connection.commit()
        except Exception as e:
            print("unidad ya ingresada ¿?")
            flash("Error al crear")

    flash("unidades importadas")
    print("unidades importadas")

def obtener_equipos():
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM equipo")
        return cur.fetchall()
    

@equipo.route('/equipos', methods=['GET'])
def listar_equipos():
    page = request.args.get('page', 1, type=int)  # Obtener el número de página de la solicitud
    per_page = 15  # Número máximo de equipos por página
    equipos = obtener_equipos()  # Función que obtiene todos los equipos
    total_equipos = len(equipos)  # Contar la cantidad total de equipos

    # Calcular el inicio y el final de los equipos para la página actual
    start = (page - 1) * per_page
    end = start + per_page
    equipos_pagina = equipos[start:end]  # Obtener solo los equipos para la página actual

    return render_template(
        'Equipo/equipo.html', 
        items=equipos_pagina, 
        total_equipos=total_equipos, 
        page=page
        )


# #buscar un equipo singular por id
@equipo.route("/buscar_equipos", methods=["GET"])
@loguear_requerido
def buscar_equipos():
    query = request.args.get("q", "").lower()  # Obtener el término de búsqueda
    page = request.args.get("page", default=1, type=int)  # Obtener la página actual
    per_page = 7  # Número de resultados por página
    offset = (page - 1) * per_page

    cur = mysql.connection.cursor()

    # Consulta para buscar equipos
    cur.execute(f"""
        SELECT *
        FROM super_equipo
        WHERE LOWER(Cod_inventarioEquipo) LIKE %s
           OR LOWER(Num_serieEquipo) LIKE %s
           OR LOWER(nombreEstado_equipo) LIKE %s
           OR LOWER(nombreFuncionario) LIKE %s
           OR LOWER(codigoproveedor_equipo) LIKE %s
           OR LOWER(nombreUnidad) LIKE %s
           OR LOWER(nombreTipo_equipo) LIKE %s
           OR LOWER(nombreModeloequipo) LIKE %s
        LIMIT %s OFFSET %s
    """, (f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%",
          f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%", per_page, offset))
    equipos = cur.fetchall()

    # Total de resultados para la búsqueda
    cur.execute(f"""
        SELECT COUNT(*) AS total
        FROM super_equipo
        WHERE LOWER(Cod_inventarioEquipo) LIKE %s
           OR LOWER(Num_serieEquipo) LIKE %s
           OR LOWER(nombreEstado_equipo) LIKE %s
           OR LOWER(nombreFuncionario) LIKE %s
           OR LOWER(codigoproveedor_equipo) LIKE %s
           OR LOWER(nombreUnidad) LIKE %s
           OR LOWER(nombreTipo_equipo) LIKE %s
           OR LOWER(nombreModeloequipo) LIKE %s
    """, (f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%",
          f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%"))
    total = cur.fetchone()["total"]
    total_pages = (total + per_page - 1) // per_page

    # Calcular las páginas visibles
    visible_pages = []
    if total_pages <= 7:  # Mostrar todas las páginas si son pocas
        visible_pages = list(range(1, total_pages + 1))
    else:
        if page > 4:
            visible_pages.append(1)
            if page > 5:
                visible_pages.append("...")
        visible_pages.extend(range(max(1, page - 2), min(total_pages + 1, page + 3)))
        if page < total_pages - 3:
            if page < total_pages - 4:
                visible_pages.append("...")
            visible_pages.append(total_pages)

    return jsonify({
        "equipos": equipos,
        "total": total,
        "total_pages": total_pages,
        "current_page": page,
        "visible_pages": visible_pages  # Enviar las páginas visibles al frontend
    })

# exportar a excel
@equipo.route("/crear_excel")
@loguear_requerido
def crear_excel():
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")

    # Obtener los IDs desde la URL
    ids_param = request.args.get("ids") 
    ids_lista = ids_param.split(",") if ids_param else None

    # Crear Excel
    wb = Workbook()
    ws = wb.active

    # Consulta SQL (si hay IDs, filtramos)
    cur = mysql.connection.cursor()
    if ids_lista:
        placeholders = ",".join(["%s"] * len(ids_lista))
        query = f"""
        SELECT * 
        FROM super_equipo se
        INNER JOIN unidad u ON u.idUnidad = se.idUnidad
        INNER JOIN modalidad m ON u.idModalidad = m.idModalidad
        INNER JOIN comuna c ON c.idComuna = u.idComuna
        INNER JOIN provincia p ON c.idProvincia = p.idProvincia
        INNER JOIN modelo_equipo me ON me.idModelo_Equipo = se.idModelo_Equipo
        INNER JOIN marca_tipo_equipo mte ON me.idMarca_Tipo_Equipo = mte.idMarcaTipo
        INNER JOIN marca_equipo mae ON mae.idMarca_Equipo = mte.idMarca_Equipo
        INNER JOIN orden_compra oc ON oc.idOrden_compra = se.idOrden_compra
        INNER JOIN proveedor pvr ON pvr.idProveedor = oc.idProveedor
        WHERE se.idEquipo IN ({placeholders})
        """
        cur.execute(query, ids_lista)
    else:
        query = """
        SELECT * 
        FROM super_equipo se
        INNER JOIN unidad u ON u.idUnidad = se.idUnidad
        INNER JOIN modalidad m ON u.idModalidad = m.idModalidad
        INNER JOIN comuna c ON c.idComuna = u.idComuna
        INNER JOIN provincia p ON c.idProvincia = p.idProvincia
        INNER JOIN modelo_equipo me ON me.idModelo_Equipo = se.idModelo_Equipo
        INNER JOIN marca_tipo_equipo mte ON me.idMarca_Tipo_Equipo = mte.idMarcaTipo
        INNER JOIN marca_equipo mae ON mae.idMarca_Equipo = mte.idMarca_Equipo
        INNER JOIN orden_compra oc ON oc.idOrden_compra = se.idOrden_compra
        INNER JOIN proveedor pvr ON pvr.idProveedor = oc.idProveedor;
        """
        cur.execute(query)

    equipo_data = cur.fetchall()

    if not equipo_data:
        flash("No hay equipos visibles para exportar.","warning")
        return redirect(request.referrer or "/")
    # generar encabezado
    # encabezado

    encabezado = [
        "Provincia", "Comuna", "Modalidad", "Código Proveedor",
        "Nombre", "Tipo de Bien", "Marca", "Modelo",
        "N° Serie", "Código Inventario", "Nombre Proveedor"
    ]
    for i in range(len(encabezado)):
        char = chr(65 + i)
        ws[char + "1"].fill = PatternFill(start_color="000ff000", fill_type="solid")
        ws.column_dimensions[char].width = 20
        ws[char + "1"] = encabezado[i]

    # Llenar los datos en el Excel
    for fila_idx, fila in enumerate(equipo_data, start=2):
        ws[f"A{fila_idx}"] = fila["nombreProvincia"]
        ws[f"B{fila_idx}"] = fila["nombreComuna"]
        ws[f"C{fila_idx}"] = fila["nombreModalidad"]
        ws[f"D{fila_idx}"] = fila["codigoproveedor_equipo"]
        ws[f"E{fila_idx}"] = fila["nombreUnidad"]
        ws[f"F{fila_idx}"] = fila["nombreTipo_equipo"]
        ws[f"G{fila_idx}"] = fila["nombreMarcaEquipo"]
        ws[f"H{fila_idx}"] = fila["nombreModeloequipo"]
        ws[f"I{fila_idx}"] = fila["Num_serieEquipo"]
        ws[f"J{fila_idx}"] = fila["Cod_inventarioEquipo"]
        ws[f"K{fila_idx}"] = fila["nombreProveedor"]

    # Guardar y devolver el archivo
    wb.save("datos_exportados.xlsx")
    return send_file("datos_exportados.xlsx", as_attachment=True)
